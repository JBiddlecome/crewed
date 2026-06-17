import calendar as _calendar
from datetime import date

from sqlalchemy.orm import Session

from . import models

FEDERAL_MIN_WAGE = 7.25

US_STATES = [
    ("AL", "Alabama"), ("AK", "Alaska"), ("AZ", "Arizona"), ("AR", "Arkansas"),
    ("CA", "California"), ("CO", "Colorado"), ("CT", "Connecticut"), ("DE", "Delaware"),
    ("DC", "District of Columbia"), ("FL", "Florida"), ("GA", "Georgia"), ("HI", "Hawaii"),
    ("ID", "Idaho"), ("IL", "Illinois"), ("IN", "Indiana"), ("IA", "Iowa"),
    ("KS", "Kansas"), ("KY", "Kentucky"), ("LA", "Louisiana"), ("ME", "Maine"),
    ("MD", "Maryland"), ("MA", "Massachusetts"), ("MI", "Michigan"), ("MN", "Minnesota"),
    ("MS", "Mississippi"), ("MO", "Missouri"), ("MT", "Montana"), ("NE", "Nebraska"),
    ("NV", "Nevada"), ("NH", "New Hampshire"), ("NJ", "New Jersey"), ("NM", "New Mexico"),
    ("NY", "New York"), ("NC", "North Carolina"), ("ND", "North Dakota"), ("OH", "Ohio"),
    ("OK", "Oklahoma"), ("OR", "Oregon"), ("PA", "Pennsylvania"), ("RI", "Rhode Island"),
    ("SC", "South Carolina"), ("SD", "South Dakota"), ("TN", "Tennessee"), ("TX", "Texas"),
    ("UT", "Utah"), ("VT", "Vermont"), ("VA", "Virginia"), ("WA", "Washington"),
    ("WV", "West Virginia"), ("WI", "Wisconsin"), ("WY", "Wyoming"),
]

STATE_NAMES = dict(US_STATES)


def get_setting(db: Session, key: str, default: str = None) -> str:
    row = db.get(models.Setting, key)
    return row.value if row else default


def set_setting(db: Session, key: str, value: str):
    row = db.get(models.Setting, key)
    if row:
        row.value = value
    else:
        db.add(models.Setting(key=key, value=value))


def effective_markup(db: Session, company: models.ClientCompany) -> float:
    """Markup percent applied on top of pay rate to produce the bill rate."""
    if company is not None and company.markup_override is not None:
        return company.markup_override
    return float(get_setting(db, "markup_percent", "55"))


def compute_bill_rate(pay_rate: float, markup_percent: float) -> float:
    return round(pay_rate * (1 + markup_percent / 100), 2)


def min_wage_for_state(db: Session, state: str) -> float:
    row = db.query(models.MinWage).filter_by(state=state).first()
    return row.rate if row else FEDERAL_MIN_WAGE


def qualifies(db: Session, user: models.User, shift: models.Shift):
    """Whether an employee can work a shift. Returns (ok, reasons)."""
    reasons = []

    # Check Block List
    blocked = (
        db.query(models.BlockList)
        .filter(
            models.BlockList.employee_id == user.id,
            models.BlockList.client_id == shift.client_id,
            (models.BlockList.location_id == None) | (models.BlockList.location_id == shift.location_id),
        )
        .first()
    )
    if blocked:
        reasons.append("You are on the block list for this client/location")

    approved_pos = {p.position_id: p for p in user.positions if p.status == "approved"}
    if shift.position_id not in approved_pos:
        all_positions = {p.position_id: p.status for p in user.positions}
        if shift.position_id in all_positions:
            status = all_positions[shift.position_id]
            reasons.append(f"{shift.position.name} is on your profile but is {status}")
        else:
            reasons.append(f"{shift.position.name} is not on your profile")
    else:
        required_level = getattr(shift, "required_level", 1) or 1
        employee_level = getattr(approved_pos[shift.position_id], "level", 2) or 2
        if employee_level < required_level:
            reasons.append(
                f"This shift requires Level {required_level} {shift.position.name} — you are Level {employee_level}"
            )

    client_position = (
        db.query(models.ClientPosition)
        .filter_by(client_id=shift.client_id, position_id=shift.position_id)
        .first()
    )
    if client_position and client_position.certs:
        today = date.today()
        held = {
            c.certification_id
            for c in user.certifications
            if c.expires_on is None or c.expires_on >= today
        }
        missing = [
            c.certification.name
            for c in client_position.certs
            if c.certification_id not in held
        ]
        if missing:
            reasons.append("Missing certification: " + ", ".join(missing))
    return (not reasons, reasons)


def refresh_shift_status(db: Session, shift: models.Shift):
    if shift.status in ("cancelled", "completed"):
        return
    db.flush()  # session has autoflush off; the count below must see pending status changes
    confirmed = (
        db.query(models.Assignment)
        .filter_by(shift_id=shift.id, status="confirmed")
        .count()
    )
    shift.status = "filled" if confirmed >= shift.headcount else "open"


def client_position_for(db: Session, shift: models.Shift):
    return (
        db.query(models.ClientPosition)
        .filter_by(client_id=shift.client_id, position_id=shift.position_id)
        .first()
    )


DETAIL_FIELDS = ("parking", "check_in_location", "check_in_contact")


def resolved_details(db: Session, shift: models.Shift) -> dict:
    """Effective parking/check-in details: shift override → event → location defaults."""
    out = {}
    for field in DETAIL_FIELDS:
        out[field] = (
            getattr(shift, field, None)
            or (getattr(shift.event, field, None) if shift.event else None)
            or getattr(shift.location, field, None)
        )
    return out


def details_map(db: Session, shifts) -> dict:
    """{shift.id: resolved details} for a list of shifts."""
    return {s.id: resolved_details(db, s) for s in shifts}


def month_weeks(year: int, month: int):
    """Calendar weeks (Sun–Sat) of date objects covering the month."""
    return _calendar.Calendar(firstweekday=6).monthdatescalendar(year, month)


def month_name(month: int) -> str:
    return _calendar.month_name[month]


def unread_count(db: Session, user: models.User) -> int:
    return (
        db.query(models.Message)
        .filter_by(recipient_id=user.id, read_at=None)
        .count()
    )


def remove_employee_from_future_shifts(db: Session, employee_id: int, client_id: int, location_id: int = None):
    """Cancel all active/requested future assignments for an employee at a client/location."""
    today = date.today()
    query = (
        db.query(models.Assignment)
        .join(models.Shift)
        .filter(
            models.Assignment.employee_id == employee_id,
            models.Shift.client_id == client_id,
            models.Shift.shift_date >= today,
            models.Assignment.status.in_(["requested", "confirmed"]),
        )
    )
    if location_id is not None:
        query = query.filter(models.Shift.location_id == location_id)

    future_assignments = query.all()
    for a in future_assignments:
        a.status = "cancelled"
        refresh_shift_status(db, a.shift)


def get_past_due_assignment(db: Session, employee_id: int):
    """Find the oldest assignment for this employee where shift_date has passed (yesterday or earlier),
    a timesheet exists, and its status is 'pending'."""
    today = date.today()
    return (
        db.query(models.Assignment)
        .join(models.Shift)
        .join(models.Timesheet)
        .filter(
            models.Assignment.employee_id == employee_id,
            models.Shift.shift_date < today,
            models.Timesheet.status == "pending",
            models.Timesheet.deleted_at == None,
        )
        .order_by(models.Shift.shift_date.asc())
        .first()
    )


def log_timesheet_event(
    db: Session,
    timesheet_id: int,
    event_type: str,
    actor_id,
    actor_role: str,
    notes: str = None,
):
    ev = models.TimesheetEvent(
        timesheet_id=timesheet_id,
        event_type=event_type,
        actor_id=actor_id,
        actor_role=actor_role,
        notes=notes,
    )
    db.add(ev)


def process_rates_excel(file_bytes: bytes) -> list:
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
            
        row_dict = dict(zip(headers, row))
        pos_name = row_dict.get('position')
        if not pos_name:
            continue
            
        pos_name = str(pos_name).strip()
        
        parsed = {'position': pos_name}
        for level in (1, 2, 3):
            pay = row_dict.get(f'level_{level}_pay')
            bill = row_dict.get(f'level_{level}_bill')
            markup = row_dict.get(f'level_{level}_markup')
            
            try:
                pay = float(pay) if pay is not None else 0.0
            except (ValueError, TypeError):
                pay = 0.0
            try:
                bill = float(bill) if bill is not None else 0.0
            except (ValueError, TypeError):
                bill = 0.0
            try:
                markup = float(markup) if markup is not None else 0.0
            except (ValueError, TypeError):
                markup = 0.0
                
            if pay > 0:
                if bill > 0 and not markup:
                    markup = ((bill - pay) / pay) * 100
                elif markup > 0 and not bill:
                    bill = pay * (1 + (markup / 100))
                elif not bill and not markup:
                    bill = pay
                    markup = 0.0
                    
            parsed[f'l{level}'] = {
                'pay': round(pay, 2),
                'bill': round(bill, 2),
                'markup': round(markup, 2)
            }
        results.append(parsed)
    return results
